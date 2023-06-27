import cv2
import os
import openpyxl
from matplotlib import pyplot as plt

# checks whether a Folder for results exists, if not creates it
def CreateFolderForResults(images_and_values, datapath_cropped = "", safe_Images_as_well = False, center_image = False):
    # Set a counter to track the number of attempts
    counter = 1
    datapath = datapath_cropped

    # Keep trying until file name doesnt exist

    while os.path.exists(datapath):
        datapath = datapath_cropped
        counter += 1
        # Append the attempt number to the path
        datapath = f"{datapath}{counter}"

    # Create the directory
    os.makedirs(datapath)
    print(f"The directory was created at " + datapath_cropped + str(counter))

    excelfile = openpyxl.Workbook()
    excelsheet = excelfile.active
    excelsheet.title = "Cover "
    excelsheet.cell(column=1, row=1, value = "Each sheet has the accuracy results for a different maskfunction")
    excelsheet.cell(column=1, row=2, value = "The rows are the images and the columns are the different configurations of the parameters")
    excelsheet.cell(column=1, row=3, value = "used in the masks")

    # could be done in one nested loop but writing consecutive values is faster
    #for idx in range(0, len(images_and_values), 1):
    #    excelsheet.cell(column=1, row=idx + 2, value=images_and_values[idx][6])


    cellCounterCol = 0
    cellCounterRow = 0
    for maskIdx in range(0, len(images_and_values[0][7]), 1):
        maskSheet = excelfile.create_sheet(title = images_and_values[0][7][maskIdx])
        mask_folder_path = os.path.join(datapath,images_and_values[0][7][maskIdx])
        maskSheet.cell(column=1, row=1, value="Folder")
        #excelsheet.cell(column=maskIdx+2, row=1, value=images_and_values[0][7][maskIdx])
        if safe_Images_as_well:
            os.makedirs(mask_folder_path)
        for idx in range(0, len(images_and_values), 1):
            maskSheet.cell(column=1, row=idx+3, value=images_and_values[idx][6])
            if safe_Images_as_well:
                new_folder_path = os.path.join(mask_folder_path, str(idx+1))
                os.makedirs(new_folder_path)
                cv2.imwrite(f"{new_folder_path}\\Original_image_" + images_and_values[idx][2] + ".jpg", images_and_values[idx][0])
            for variationIdx in range (0, len(images_and_values[idx][8][maskIdx]), 1):
                word = ""
                filename = ""
                for wordpart in range (0, len(images_and_values[idx][8][maskIdx][variationIdx]),1):
                    word = word + str(images_and_values[idx][8][maskIdx][variationIdx][wordpart]) +" "
                for wordpart in range(1, len(images_and_values[idx][8][maskIdx][variationIdx]), 2):
                    filename = "_" + filename + str(images_and_values[idx][8][maskIdx][variationIdx][wordpart])

                maskSheet.cell(column=(variationIdx*2)+2, row=1, value=word)
                maskSheet.cell(column=(variationIdx * 2) + 2, row=2, value="Accuracy")
                maskSheet.cell(column=(variationIdx * 2) + 3, row=2, value="IoU")
                maskSheet.cell(column=(variationIdx*2)+2, row=idx+3, value = images_and_values[idx][5][maskIdx][variationIdx])
                maskSheet.cell(column=(variationIdx*2)+3, row=idx+3, value = images_and_values[idx][9][maskIdx][variationIdx])
                if safe_Images_as_well:
                    cv2.imwrite(f"{new_folder_path}\\Cropped_"+images_and_values[idx][2]+ filename+".jpg", images_and_values[idx][1][maskIdx][variationIdx])
                    cv2.imwrite(f"{new_folder_path}\\Mask_" + images_and_values[idx][2]+filename+".jpg", images_and_values[idx][3][maskIdx][variationIdx])
                    if center_image:
                        cv2.imwrite(f"{new_folder_path}\\Centered_" + images_and_values[idx][2]+filename+".jpg", images_and_values[idx][4][maskIdx][variationIdx])

    print(os.path.join(datapath,"Results.xlsx"))
    excelfile.save(f'{os.path.join(datapath,"Results.xlsx")}')
    print(f"The directory was filled.")